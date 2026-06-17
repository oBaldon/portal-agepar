from __future__ import annotations

import io
from datetime import datetime, time, timedelta
from typing import Any, Dict, Iterable, Optional
from zoneinfo import ZoneInfo

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.automations import tasks as tasks_automation
from app.db import _pg

REPORT_TZ = ZoneInfo("America/Sao_Paulo")
NO_ROLE_SHEET_KEY = "__sem_cargo__"
BUSINESS_WEEK_LABEL = "segunda a sexta"

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

STATUS_FILLS = {
    "A fazer": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "Em andamento": PatternFill(fill_type="solid", fgColor="D9EAF7"),
    "Em revisão": PatternFill(fill_type="solid", fgColor="EADCF8"),
    "Concluída": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    "Cancelada": PatternFill(fill_type="solid", fgColor="F4CCCC"),
}

PRIORITY_FILLS = {
    "Alta": PatternFill(fill_type="solid", fgColor="FCE5CD"),
    "Média": PatternFill(fill_type="solid", fgColor="EDEDED"),
    "Baixa": PatternFill(fill_type="solid", fgColor="D9EAD3"),
}

COLUMN_WIDTHS = {
    "A": 24,
    "B": 30,
    "C": 52,
    "D": 18,
    "E": 14,
    "F": 22,
    "G": 24,
    "H": 18,
    "I": 18,
    "J": 18,
    "K": 14,
    "L": 19,
}



def _norm_roles(user: Dict[str, Any]) -> set[str]:
    return {str(role).strip().lower() for role in (user.get("roles") or []) if str(role).strip()}


def _directory_roles() -> list[str]:
    return [str(role).strip().lower() for role in tasks_automation._load_role_options() if str(role).strip()]


def _is_full_export_user(user: Dict[str, Any]) -> bool:
    roles = _norm_roles(user)
    return user.get("is_superuser") is True or "admin" in roles


def _report_scope_roles(user: Dict[str, Any]) -> list[str]:
    user_roles = _norm_roles(user)
    return [role for role in _directory_roles() if role in user_roles]


def _week_window(reference: Optional[datetime] = None) -> tuple[datetime, datetime]:
    now_local = (reference.astimezone(REPORT_TZ) if reference else datetime.now(REPORT_TZ))
    monday = now_local.date() - timedelta(days=now_local.weekday())
    start = datetime.combine(monday, time.min, tzinfo=REPORT_TZ)
    # Regra de negócio consolidada: o compilado semanal considera a semana útil,
    # de segunda a sexta. O limite superior segue exclusivo para simplificar os filtros SQL.
    end = start + timedelta(days=5)
    return start, end


def _sheet_label(role_name: Optional[str]) -> str:
    role = str(role_name or "").strip().lower()
    if not role:
        return "SEM_CARGO"
    if role == NO_ROLE_SHEET_KEY:
        return "SEM_CARGO"
    return role.upper()


def _scope_label(scope: str, role_names: Iterable[str]) -> str:
    if scope == "full":
        return "Todos os cargos"
    labels = [_sheet_label(role) for role in role_names]
    if not labels:
        return "Indisponível"
    if len(labels) == 1:
        return labels[0]
    return ", ".join(labels)


def build_weekly_export_context(user: Dict[str, Any]) -> Dict[str, Any]:
    week_start, week_end = _week_window()
    role_names = _report_scope_roles(user)

    if _is_full_export_user(user):
        scope = "full"
        sheet_keys = [*_directory_roles(), NO_ROLE_SHEET_KEY]
        enabled = True
    elif role_names:
        scope = "role_scoped"
        sheet_keys = role_names
        enabled = True
    else:
        scope = "unavailable"
        sheet_keys = []
        enabled = False

    return {
        "enabled": enabled,
        "scope": scope,
        "scopeLabel": _scope_label(scope, role_names),
        "roleNames": role_names,
        "sheetKeys": sheet_keys,
        "sheetLabels": [_sheet_label(key) for key in sheet_keys],
        "weekStart": week_start.date().isoformat(),
        "weekEnd": (week_end - timedelta(microseconds=1)).date().isoformat(),
        "weekLabel": f"{week_start.date().isoformat()} a {(week_end - timedelta(microseconds=1)).date().isoformat()}",
        "weekCadenceLabel": BUSINESS_WEEK_LABEL,
        "timezone": "America/Sao_Paulo",
    }


def _load_weekly_rows(
    *,
    week_start: datetime,
    week_end: datetime,
    scope: str,
    role_names: list[str],
) -> list[dict[str, Any]]:
    where_clauses = [
        "t.deleted_at IS NULL",
        "("
        "  COALESCE(prev.status_before_week = 'em_andamento', FALSE)"
        "  OR evt.started_in_week_at IS NOT NULL"
        "  OR evt.completed_in_week_at IS NOT NULL"
        "  OR evt.cancelled_in_week_at IS NOT NULL"
        ")",
    ]
    params: list[Any] = []

    if scope != "full":
        if not role_names:
            raise HTTPException(status_code=403, detail="weekly export is not available for this user")
        placeholders = ", ".join(["%s"] * len(role_names))
        where_clauses.append(f"COALESCE(t.assigned_role_name, '') IN ({placeholders})")
        params.extend(role_names)

    where_sql = " AND ".join(where_clauses)

    with _pg() as conn, conn.cursor() as cur:
        cur.execute(
            f"""
            SELECT
              t.id,
              t.title,
              t.description,
              t.status,
              t.priority,
              t.due_date,
              t.created_at,
              t.updated_at,
              COALESCE(NULLIF(BTRIM(t.assigned_role_name), ''), NULL) AS assigned_role_name,
              creator.name AS created_by_name,
              assignee.name AS assigned_to_name,
              prev.status_before_week,
              evt.started_in_week_at,
              evt.completed_in_week_at,
              evt.cancelled_in_week_at
            FROM tasks t
            LEFT JOIN users creator ON creator.id = t.created_by_user_id
            LEFT JOIN users assignee ON assignee.id = t.assigned_to_user_id
            LEFT JOIN (
              SELECT DISTINCT ON (e.task_id)
                e.task_id,
                COALESCE(NULLIF(BTRIM(e.new_value->>'status'), ''), NULL) AS status_before_week
              FROM task_events e
              WHERE e.created_at < %s
                AND e.event_type IN ('task_created', 'task_status_changed')
                AND COALESCE(NULLIF(BTRIM(e.new_value->>'status'), ''), NULL) IS NOT NULL
              ORDER BY e.task_id ASC, e.created_at DESC, e.id DESC
            ) prev ON prev.task_id = t.id
            LEFT JOIN (
              SELECT
                e.task_id,
                MIN(
                  CASE
                    WHEN e.event_type IN ('task_created', 'task_status_changed')
                     AND COALESCE(e.new_value->>'status', '') = 'em_andamento'
                    THEN e.created_at
                    ELSE NULL
                  END
                ) AS started_in_week_at,
                MIN(
                  CASE
                    WHEN e.event_type = 'task_completed'
                      OR (
                        e.event_type = 'task_status_changed'
                        AND COALESCE(e.new_value->>'status', '') = 'concluida'
                      )
                    THEN e.created_at
                    ELSE NULL
                  END
                ) AS completed_in_week_at,
                MIN(
                  CASE
                    WHEN e.event_type = 'task_cancelled'
                      OR (
                        e.event_type = 'task_status_changed'
                        AND COALESCE(e.new_value->>'status', '') = 'cancelada'
                      )
                    THEN e.created_at
                    ELSE NULL
                  END
                ) AS cancelled_in_week_at
              FROM task_events e
              WHERE e.created_at >= %s
                AND e.created_at < %s
                AND (
                  (
                    e.event_type IN ('task_created', 'task_status_changed')
                    AND COALESCE(e.new_value->>'status', '') = 'em_andamento'
                  )
                  OR e.event_type IN ('task_completed', 'task_cancelled')
                  OR (
                    e.event_type = 'task_status_changed'
                    AND COALESCE(e.new_value->>'status', '') IN ('concluida', 'cancelada')
                  )
                )
              GROUP BY e.task_id
            ) evt ON evt.task_id = t.id
            WHERE {where_sql}
            ORDER BY
              COALESCE(NULLIF(BTRIM(t.assigned_role_name), ''), 'zzzz') ASC,
              COALESCE(assignee.name, 'Sem responsável') ASC,
              COALESCE(evt.completed_in_week_at, evt.cancelled_in_week_at, evt.started_in_week_at, t.updated_at, t.created_at) DESC,
              t.title ASC
            """,
            [
                week_start,
                week_start,
                week_end,
                *params,
            ],
        )
        rows = cur.fetchall() or []

    return [dict(row) for row in rows]


def _localize_datetime(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        try:
            localized = value.astimezone(REPORT_TZ)
        except ValueError:
            localized = value
        return localized.strftime("%d/%m/%Y %H:%M")
    return str(value)


def _localize_date(value: Any) -> str:
    if value is None:
        return ""
    try:
        return value.strftime("%d/%m/%Y")
    except Exception:
        return str(value)



def _display_text(value: Any, fallback: str = "—") -> str:
    text = " ".join(str(value or "").split())
    return text or fallback


def _friendly_week_activity(row: dict[str, Any]) -> str:
    if row.get("started_in_week_at") is not None:
        return "Iniciada nesta semana"

    status_before_week = str(row.get("status_before_week") or "").strip().lower()
    if status_before_week == "em_andamento":
        return "Já estava em andamento"

    return "—"


def _description_preview(value: Any, max_len: int = 120) -> str:
    text = " ".join(str(value or "").split())
    if len(text) <= max_len:
        return text
    return text[: max_len - 1].rstrip() + "…"


def _group_rows_by_sheet(
    rows: list[dict[str, Any]],
    *,
    scope: str,
    role_names: list[str],
) -> dict[str, list[dict[str, Any]]]:
    if scope == "full":
        buckets: dict[str, list[dict[str, Any]]] = {role: [] for role in _directory_roles()}
        buckets[NO_ROLE_SHEET_KEY] = []
    else:
        buckets = {role: [] for role in role_names}

    for row in rows:
        key = str(row.get("assigned_role_name") or "").strip().lower() or NO_ROLE_SHEET_KEY
        if key not in buckets:
            continue
        buckets[key].append(row)
    return buckets


def _sheet_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    output: list[list[Any]] = []
    for row in rows:
        output.append(
            [
                _display_text(row.get("assigned_to_name"), "Sem responsável"),
                _display_text(row.get("title"), "Sem título"),
                _description_preview(row.get("description"), max_len=280) or "Sem descrição informada",
                tasks_automation._status_label(row.get("status")),
                str(row.get("priority") or "").capitalize() or "—",
                _display_text(row.get("created_by_name")),
                _friendly_week_activity(row),
                _localize_datetime(row.get("started_in_week_at")) or "—",
                _localize_datetime(row.get("completed_in_week_at")) or "—",
                _localize_datetime(row.get("cancelled_in_week_at")) or "—",
                _localize_date(row.get("due_date")) or "—",
                _localize_datetime(row.get("updated_at")) or "—",
            ]
        )
    return output


def _autosize_columns(ws) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 38)


def _apply_column_layout(ws) -> None:
    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width


def _style_data_rows(ws, *, start_row: int, end_row: int) -> None:
    striped_fill = PatternFill(fill_type="solid", fgColor="FAFAFA")
    for row_idx in range(start_row, end_row + 1):
        if row_idx % 2 == 1:
            for cell in ws[row_idx]:
                cell.fill = striped_fill

        ws.row_dimensions[row_idx].height = 36

        status_cell = ws[f"D{row_idx}"]
        priority_cell = ws[f"E{row_idx}"]

        status_fill = STATUS_FILLS.get(str(status_cell.value or "").strip())
        if status_fill is not None:
            status_cell.fill = status_fill

        priority_fill = PRIORITY_FILLS.get(str(priority_cell.value or "").strip())
        if priority_fill is not None:
            priority_cell.fill = priority_fill

        for cell in ws[row_idx]:
            cell.border = THIN_BORDER


def _write_sheet_intro(ws, *, label: str, context: dict[str, Any], headers_len: int, total_tasks: int) -> int:
    last_col = get_column_letter(headers_len)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"Compilado semanal de tarefas — {label} — {context['weekLabel']}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(vertical="center", wrap_text=True)

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        "Leitura rápida: este relatório mostra as atividades que já estavam em andamento, "
        "começaram, foram concluídas ou canceladas na semana útil selecionada."
    )
    ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = f"Total de atividades nesta aba: {total_tasks}"
    ws["A3"].font = Font(italic=True)
    ws["A3"].alignment = Alignment(vertical="center")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 20
    return 4


def _build_workbook(
    *,
    grouped_rows: dict[str, list[dict[str, Any]]],
    context: dict[str, Any],
) -> bytes:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    headers = [
        "Responsável",
        "Atividade",
        "Descrição da Atividade",
        "Situação Atual",
        "Prioridade",
        "Solicitada por",
        "Andamento na Semana",
        "Início na Semana",
        "Conclusão na Semana",
        "Cancelamento na Semana",
        "Prazo",
        "Última Atualização",
    ]

    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    header_font = Font(bold=True)
    top_alignment = Alignment(vertical="top", wrap_text=True)

    for sheet_key in context["sheetKeys"]:
        label = _sheet_label(sheet_key)
        ws = wb.create_sheet(title=label[:31])
        header_row = _write_sheet_intro(
            ws,
            label=label,
            context=context,
            headers_len=len(headers),
            total_tasks=len(grouped_rows.get(sheet_key, [])),
        )

        ws.append(headers)
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = top_alignment
            cell.border = THIN_BORDER

        data_rows = _sheet_rows(grouped_rows.get(sheet_key, []))
        for values in data_rows:
            ws.append(values)

        ws.freeze_panes = f"A{header_row + 1}"

        if ws.max_row == header_row:
            ws.append(
                [
                    "Nenhuma atividade desta aba esteve em andamento, foi iniciada, concluída ou cancelada na semana selecionada.",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )

        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(ws.max_row, header_row)}"

        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
            for cell in row:
                cell.alignment = top_alignment

        _autosize_columns(ws)
        _apply_column_layout(ws)
        _style_data_rows(ws, start_row=header_row + 1, end_row=ws.max_row)

    wb.properties.creator = "Portal AGEPAR"
    wb.properties.title = f"Compilado semanal de tarefas ({context['weekLabel']})"
    wb.properties.description = (
        "Compilado semanal de tarefas por aba/cargo, com linguagem simplificada e foco nas atividades "
        "que estiveram em andamento ou tiveram movimentação na semana útil (segunda a sexta, horário de Brasília)."
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_filename(context: dict[str, Any]) -> str:
    week_start = context["weekStart"]
    week_end = context["weekEnd"]

    if context["scope"] == "full":
        scope_suffix = "geral"
    elif len(context["sheetKeys"]) == 1:
        scope_suffix = _sheet_label(context["sheetKeys"][0]).lower()
    else:
        scope_suffix = "multiplos"

    return f"agepar_gestao_tarefas_compilado_semanal_{scope_suffix}_{week_start}_a_{week_end}.xlsx"


def generate_weekly_task_report(user: Dict[str, Any]) -> tuple[bytes, str, Dict[str, Any]]:
    context = build_weekly_export_context(user)
    if not context["enabled"]:
        raise HTTPException(status_code=403, detail="weekly export is not available for this user")

    week_start, week_end = _week_window()
    rows = _load_weekly_rows(
        week_start=week_start,
        week_end=week_end,
        scope=context["scope"],
        role_names=context["roleNames"],
    )
    grouped_rows = _group_rows_by_sheet(
        rows,
        scope=context["scope"],
        role_names=context["roleNames"],
    )
    content = _build_workbook(grouped_rows=grouped_rows, context=context)
    context["sheetTaskCounts"] = _sheet_task_counts(grouped_rows)
    filename = _build_filename(context)
    return content, filename, context


def build_weekly_export_context_for_role(role_name: str) -> Dict[str, Any]:
    normalized_role = str(role_name or "").strip().lower()
    if normalized_role not in _directory_roles():
        raise HTTPException(status_code=422, detail={"roleName": f"invalid role (use {', '.join(_directory_roles())})"})

    week_start, week_end = _week_window()
    return {
        "enabled": True,
        "scope": "role_scoped",
        "scopeLabel": _sheet_label(normalized_role),
        "roleNames": [normalized_role],
        "sheetKeys": [normalized_role],
        "sheetLabels": [_sheet_label(normalized_role)],
        "weekStart": week_start.date().isoformat(),
        "weekEnd": (week_end - timedelta(microseconds=1)).date().isoformat(),
        "weekLabel": f"{week_start.date().isoformat()} a {(week_end - timedelta(microseconds=1)).date().isoformat()}",
        "weekCadenceLabel": BUSINESS_WEEK_LABEL,
        "timezone": "America/Sao_Paulo",
    }


def _sheet_task_counts(grouped_rows: dict[str, list[dict[str, Any]]]) -> dict[str, int]:
    return {str(key): len(value or []) for key, value in grouped_rows.items()}


def generate_weekly_task_report_for_role(role_name: str) -> tuple[bytes, str, Dict[str, Any]]:
    context = build_weekly_export_context_for_role(role_name)
    week_start, week_end = _week_window()
    rows = _load_weekly_rows(
        week_start=week_start,
        week_end=week_end,
        scope=context["scope"],
        role_names=context["roleNames"],
    )
    grouped_rows = _group_rows_by_sheet(
        rows,
        scope=context["scope"],
        role_names=context["roleNames"],
    )
    content = _build_workbook(grouped_rows=grouped_rows, context=context)
    context["sheetTaskCounts"] = _sheet_task_counts(grouped_rows)
    filename = _build_filename(context)
    return content, filename, context
