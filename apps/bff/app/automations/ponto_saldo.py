# apps/bff/app/automations/ponto_saldo.py
"""
Automação "ponto_saldo": lê espelho de ponto em PDF (mês), calcula:
- esperado (8h em dia útil; 0 em fim de semana; 0 se o PDF marcar FERIADO/PONTO FACULTATIVO),
- creditado (prioriza coluna "Resultado"; fallback Trabalhadas+Justificadas;
  se só houver texto de ATESTADO/LICENÇA sem HH:MM → credita 8h),
- faltante total / sobrante total / saldo líquido,
- plano uniforme (a partir de HOJE) para zerar o mês, sem arredondar minutos.

Importante
----------
- Universo = PDF: não existe calendário externo.
- Não persistimos o PDF (decisão de privacidade). Apenas metadata e resultado.
- Para extrair texto do PDF, este módulo usa `pypdf` (precisa estar no requirements).
"""

from __future__ import annotations

import json
import logging
import re
import unicodedata
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional, Tuple
from uuid import uuid4
from zoneinfo import ZoneInfo

from fastapi import APIRouter, BackgroundTasks, File, HTTPException, Request, UploadFile
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, ConfigDict, Field
from starlette.responses import HTMLResponse, StreamingResponse

from app.db import add_audit, get_submission, insert_submission, list_submissions, update_submission

logger = logging.getLogger(__name__)

KIND = "ponto_saldo"
PONTO_SALDO_VERSION = "0.1.0"

router = APIRouter(
    prefix="/api/automations/ponto_saldo",
    tags=["automations", "ponto_saldo"],
)

templates = Jinja2Templates(directory=str(Path(__file__).parent / "templates"))

TZ_SP = ZoneInfo("America/Sao_Paulo")
MINUTES_PER_BUSINESS_DAY = 480  # 08:00


# ----------------------------
# Helpers: tempo HH:MM
# ----------------------------
_HHMM_TOKEN_RE = re.compile(r"\b\d{1,3}:\d{2}\b")

def hhmm_to_minutes(value: Optional[str]) -> Optional[int]:
    if not value:
        return None
    s = value.strip()
    neg = s.startswith("-")
    if neg:
        s = s[1:]
    m = re.search(r"\b(\d{1,2}):(\d{2})\b", s)
    if not m:
        return None
    hh = int(m.group(1))
    mm = int(m.group(2))
    if hh < 0 or not (0 <= mm <= 59):
        return None
    total = hh * 60 + mm
    return -total if neg else total


def minutes_to_hhmm(value: int) -> str:
    sign = "-" if value < 0 else ""
    value = abs(value)
    hh = value // 60
    mm = value % 60
    return f"{sign}{hh:02d}:{mm:02d}"
 
def _xlsx_from_result(result: Dict[str, Any]) -> bytes:
    """
    Gera um XLSX com 3 abas: Resumo, Plano, Dias.
    """
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Font, Alignment  # type: ignore
    except Exception as e:
        raise HTTPException(status_code=500, detail="Dependência ausente: instale openpyxl no BFF.") from e

    wb = Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    meta = result.get("meta") or {}
    totals = result.get("totals") or {}
    planner = result.get("planner") or {}
    schedule = planner.get("schedule") or []
    days = result.get("days") or []

    def fmt_signed_minutes(v: int) -> str:
        try:
            n = int(v)
        except Exception:
            n = 0
        return ("+" if n >= 0 else "-") + minutes_to_hhmm(abs(n))


    # --- Aba: Resumo ---
    ws = wb.create_sheet("Resumo")
    ws.append(["Campo", "Valor"])
    ws["A1"].font = bold
    ws["B1"].font = bold

    rows = [
        ("Competência", meta.get("period", "")),
        ("Gerado em", meta.get("generated_at", "")),
        ("Esperado (mês)", minutes_to_hhmm(int(totals.get("expected_minutes") or 0))),
        ("Creditado (até agora)", minutes_to_hhmm(int(totals.get("credited_minutes") or 0))),
        ("Faltando p/ fechar", minutes_to_hhmm(int(totals.get("missing_minutes_total") or 0))),
        ("Sobrando p/ fechar", minutes_to_hhmm(int(totals.get("extra_minutes_total") or 0))),
        ("Saldo acumulado (até agora)", minutes_to_hhmm(int(totals.get("so_far_net_balance_minutes") or 0))),
        ("Dias úteis restantes", int(planner.get("available_business_days") or 0)),
    ]
    for k, v in rows:
        ws.append([k, v])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 44
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for c in r:
            c.alignment = left

    # --- Aba: Plano ---
    ws = wb.create_sheet("Plano")
    ws.append(["Data", "Total sugerido", "Ajuste", "Observação"])
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = left

    if isinstance(schedule, list):
        for it in schedule:
            total = int((it or {}).get("suggested_total_for_day_minutes") or 0)
            adj = int((it or {}).get("suggested_adjust_minutes") or 0)
            ws.append([
                (it or {}).get("date", ""),
                minutes_to_hhmm(total),
                minutes_to_hhmm(adj),
                (it or {}).get("comment", "") or "",
            ])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 64
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for c in r:
            c.alignment = left

    # --- Aba: Dias ---
    ws = wb.create_sheet("Dias")
    ws.append(["Data", "Tipo", "Esperado", "Creditado", "Saldo", "Marcações", "Status"])
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = left

    if isinstance(days, list):
        for d in days:
            d = d or {}
            markers = d.get("markers_from_pdf") or []
            issues = d.get("issues") or []
            day_type = (d.get("day_type") or "").strip()

            # Para ficar igual ao exemplo: não listar fins de semana
            if day_type == "WEEKEND":
                continue

            dt = (d.get("date") or "").strip()
            wd = (d.get("weekday") or "").strip()
            data_label = f"{dt} ({wd})" if (dt and wd) else (dt or "")

            exp = int(d.get("expected_minutes") or 0)
            cred = int(d.get("credited_minutes") or 0)
            bal = int(d.get("balance_minutes") or 0)

            markers_txt = ", ".join([str(x) for x in markers]) if markers else "—"
            if not issues:
                status_txt = "OK"
            elif "MISSING_DATA" in issues:
                status_txt = "FALTANDO DADOS"
            elif "PENDING" in issues:
                status_txt = "PENDENTE"
            else:
                status_txt = ", ".join([str(x) for x in issues])

            ws.append([
                data_label,
                day_type,
                minutes_to_hhmm(exp),
                minutes_to_hhmm(cred),
                fmt_signed_minutes(bal),
                markers_txt,
                status_txt,
            ])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 18  # 2026-02-02 (MON)
    ws.column_dimensions["B"].width = 18  # BUSINESS_DAY / HOLIDAY / PF
    ws.column_dimensions["C"].width = 10  # 08:00
    ws.column_dimensions["D"].width = 10  # 07:52
    ws.column_dimensions["E"].width = 10  # +00:00 / -00:08
    ws.column_dimensions["F"].width = 28  # marcações
    ws.column_dimensions["G"].width = 16  # OK / PENDENTE / FALTANDO DADOS
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
        for c in r:
            c.alignment = left

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _current_user(req: Request) -> Optional[Dict[str, Any]]:
    try:
        return req.session.get("user")
    except Exception:
        return None


# ----------------------------
# PDF text extraction (pypdf)
# ----------------------------
def normalize_pdf_text(text: str) -> str:
    """
    Normaliza quebras comuns do extract_text():
    - junta dígitos quando um deles está sozinho numa linha e o próximo inicia data/hora (ex.: "1\n0/02" => "10/02", "1\n3:03" => "13:03")
    Mantém a segurança: não junta finais de tokens (ex.: "08:00\n03/02" não é afetado).
    """
    t = text or ""
    for _ in range(8):
        new = re.sub(r"(?m)^\s*(\d)\s*\n\s*(\d)(?=[:/])", r"\1\2", t)
        if new == t:
            break
        t = new
    return t


def extract_text_from_pdf(pdf_bytes: bytes) -> str:
    """
    Extrai texto do PDF usando pypdf. Funciona melhor para PDFs "texto-selecionável".
    Se páginas vierem sem texto, insere um marcador para permitir insight/alerta.
    """
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail="Dependência ausente para ler PDF. Instale 'pypdf' no BFF.",
        ) from e

    reader = PdfReader(BytesIO(pdf_bytes))
    parts: List[str] = []
    for i, page in enumerate(reader.pages):
        try:
            t = page.extract_text() or ""
        except Exception:
            t = ""
        if t.strip():
            parts.append(normalize_pdf_text(t))
        else:
            parts.append(f"\n[PAGE_{i}_NO_TEXT]\n")
    return "\n".join(parts)

# ----------------------------
# Normalização e helpers de parsing
# ----------------------------
MONTHS_PT = {
    "JANEIRO": 1, "FEVEREIRO": 2, "MARCO": 3, "MARÇO": 3, "ABRIL": 4, "MAIO": 5, "JUNHO": 6,
    "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
}

def _norm_no_space_upper(s: str) -> str:
    s = s or ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")  # remove acentos
    s = re.sub(r"\s+", "", s).upper()
    return s

def _cut_footer_from_block(block: str) -> str:
    """
    Evita que o rodapé (Semana/Legenda/timestamp) "contamine" o último dia.
    """
    up = block.upper()
    cut_points = []
    for key in ("SEMANA 1:", "SEMANA 1", "LEGENDA:", "LEGENDA", "FUNCIONÁRIO CADASTRADO", "FUNCIONARIO CADASTRADO"):
        p = up.find(key)
        if p >= 0:
            cut_points.append(p)
    if cut_points:
        return block[: min(cut_points)]
    return block

# ----------------------------
# Parsing do espelho de ponto
# ----------------------------
DATE_LINE_RE = re.compile(r"(?P<dd>\d{1,2})/(?P<mm>\d{1,2})\s*-\s*(?P<weekday>[A-Za-zÀ-ÿ]+)", re.IGNORECASE)

KEY_PF = "PONTO FACULTATIVO"
KEY_FERIADO = "FERIADO"
KEY_ATESTADO = "ATESTADO"
KEY_LICENCA = "LICENÇA"  # pode vir sem acento no texto extraído
KEY_FERIAS = "FERIAS"    # no PDF costuma vir sem acento

MONTHS_PT = {
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARCO": 3,
    "MARÇO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
    "DEZEMBRO": 12,
}


def _norm_no_space_upper(s: str) -> str:
    s = s or ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")  # remove acentos
    s = re.sub(r"\s+", "", s).upper()
    return s


def detect_period(text: str) -> Tuple[int, int]:
    """
    Tenta detectar mês/ano do PDF.
    Estratégia:
    - procurar mm/yyyy no documento (muitos relatórios trazem no cabeçalho/rodapé)
    - procurar "Fevereiro/2026" (mês por extenso)
    - se não achar, usa mês do primeiro dd/mm e ano atual (SP).
    """
    # Ex.: Fevereiro/2026
    m2 = re.search(r"(?i)\b([A-Za-zÀ-ÿ]+)\s*/\s*(\d{4})\b", text)
    if m2:
        mon_txt = _norm_no_space_upper(m2.group(1))
        yyyy = int(m2.group(2))
        if mon_txt in MONTHS_PT:
            return MONTHS_PT[mon_txt], yyyy

    m = re.search(r"\b(?P<mm>\d{2})/(?P<yyyy>\d{4})\b", text)
    if m:
        return int(m.group("mm")), int(m.group("yyyy"))

    first = DATE_LINE_RE.search(text)
    now_year = datetime.now(TZ_SP).date().year
    if first:
        return int(first.group("mm")), now_year

    now = datetime.now(TZ_SP).date()
    return now.month, now.year


def parse_days(text: str) -> List[Dict[str, Any]]:
    """
    Retorna lista de dias a partir do texto do PDF.
    Cada "dia" é o bloco entre uma linha dd/mm - ... e a próxima.
    """
    matches = list(DATE_LINE_RE.finditer(text))
    if not matches:
        return []

    _, year = detect_period(text)

    out: List[Dict[str, Any]] = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        block = _cut_footer_from_block(text[start:end])

        dd = int(m.group("dd"))
        mm = int(m.group("mm"))
        d = date(year, mm, dd)

        compact = _norm_no_space_upper(block)

        markers: List[str] = []
        if _norm_no_space_upper(KEY_PF) in compact:
            markers.append(KEY_PF)
        if _norm_no_space_upper(KEY_FERIADO) in compact:
            markers.append(KEY_FERIADO)
        if _norm_no_space_upper(KEY_ATESTADO) in compact:
            markers.append(KEY_ATESTADO)
        # "LICENÇA" pode virar "LICENCA" sem acento; usamos versão sem acento
        if "LICENCA" in compact or "LICEN" in compact:
            markers.append(KEY_LICENCA)

        # FÉRIAS pode vir como "FERIAS" (sem acento) e às vezes quebrado por linha
        if "FERIAS" in compact:
            markers.append(KEY_FERIAS)

        # Captura todos HH:MM (inclui negativos caso existam).
        hhmm_tokens = _HHMM_TOKEN_RE.findall(block)
        token_count = len(hhmm_tokens)

        worked = justified = result = ""
        # Regra prática para este modelo de espelho:
        # - Dia "completo" normalmente tem >= 5 horários (4 batidas + total do dia).
        # - Se tiver só 1 horário e NÃO houver justificativa (ex.: hoje só com entrada), não é total.
        if token_count >= 5:
            result = hhmm_tokens[-1]
        elif token_count == 1 and (KEY_ATESTADO in markers):
            # Atestado geralmente aparece só com o total "08:00" na coluna de justificativa.
            result = hhmm_tokens[0]
        else:
            result = ""
        out.append(
            {
                "date": d.isoformat(),
                "weekday": d.strftime("%a").upper(),
                "markers_from_pdf": markers,
                "time_tokens_count": token_count,
                "pdf_values": {
                    "worked_hhmm": worked,
                    "justified_hhmm": justified,
                    "result_hhmm": result,
                },
            }
        )

    return out


# ----------------------------
# Regras do cálculo
# ----------------------------
DayType = Literal["BUSINESS_DAY", "WEEKEND", "HOLIDAY", "PONTO_FACULTATIVO"]
CreditedSource = Literal["RESULT", "WORKED_PLUS_JUSTIFIED", "TEXT_ONLY_FULL_DAY", "LEAVE_DAY", "ZERO"]


def compute_expected_minutes(day_iso: str, markers: List[str]) -> Tuple[int, DayType]:
    d = date.fromisoformat(day_iso)
    if d.weekday() >= 5:
        return 0, "WEEKEND"
    # FÉRIAS: não conta como dia esperado no mês (expected=0)
    if KEY_FERIAS in markers:
        return 0, "HOLIDAY"
    # LICENÇA: não conta como dia esperado (no PDF/PSF esse tipo não entra no total do mês)
    if KEY_LICENCA in markers:
        return 0, "HOLIDAY"
    if KEY_PF in markers:
        return 0, "PONTO_FACULTATIVO"
    if KEY_FERIADO in markers:
        return 0, "HOLIDAY"
    return MINUTES_PER_BUSINESS_DAY, "BUSINESS_DAY"


def compute_credited_minutes(
    pdf_values: Dict[str, str],
    markers: List[str],
) -> Tuple[int, CreditedSource, List[str]]:
    notes: List[str] = []

    # FÉRIAS: não credita horas; fica neutralizado via expected=0
    if KEY_FERIAS in markers:
        notes.append("Dia de FÉRIAS: expected=0 e credited=0 (não entra no total mensal).")
        return 0, "LEAVE_DAY", notes

    # LICENÇA: não credita horas automaticamente (neutralizado via expected=0)
    if KEY_LICENCA in markers:
        notes.append("Dia de LICENÇA: não entra no total mensal (expected=0, credited=0).")
        return 0, "LEAVE_DAY", notes


    # 1) Resultado é fonte principal
    res = hhmm_to_minutes(pdf_values.get("result_hhmm") or "")
    if res is not None:
        return res, "RESULT", notes

    # 2) Fallback: trabalhadas + justificadas
    w = hhmm_to_minutes(pdf_values.get("worked_hhmm") or "") or 0
    j = hhmm_to_minutes(pdf_values.get("justified_hhmm") or "") or 0
    if w > 0 or j > 0:
        return (w + j), "WORKED_PLUS_JUSTIFIED", notes

    # 3) Texto-only leave: ATESTADO/LICENÇA sem números → 8h creditadas
    if KEY_ATESTADO in markers:
        notes.append("Texto de ATESTADO sem horas numéricas: creditado dia integral (08:00).")
        return MINUTES_PER_BUSINESS_DAY, "TEXT_ONLY_FULL_DAY", notes

    # 4) Sem dados
    return 0, "ZERO", notes


def build_planner(
    days: List[Dict[str, Any]],
    required_total_remaining_minutes: int,
    recommended_max_adjust_minutes_per_day: int = 120,
) -> Dict[str, Any]:
    """
    Plano uniforme a partir de HOJE (SP), sem arredondar minutos.
    Aqui o "required_total_remaining_minutes" é o TOTAL (em minutos) que precisa ser feito
    nos dias úteis restantes para fechar o mês.
    """
    today = datetime.now(TZ_SP).date()

    # dias disponíveis: >= hoje, expected=480 (dia útil).
    # Se hoje já tem total (não está PENDING), não entra no planner.
    available: List[date] = []
    for d in days:
        di = date.fromisoformat(d["date"])
        if di < today:
            continue
        if d.get("expected_minutes") == MINUTES_PER_BUSINESS_DAY:
            if di == today:
                issues = d.get("issues") or []
                if "PENDING" not in issues:
                    continue
            available.append(di)

    n = len(available)
    schedule: List[Dict[str, Any]] = []
    warnings: List[str] = []
    hm_warnings: List[str] = []

    if n == 0:
        if required_total_remaining_minutes != 0:
            warnings.append("Não há dias úteis disponíveis a partir de hoje dentro deste mês para distribuir o saldo.")
        return {
            "strategy": "UNIFORM",
            "available_business_days": 0,
            "target": "ZERO_OUT_MONTH",
            "schedule": [],
            "healthy_mode": {
                "enabled": True,
                "recommended_max_adjust_minutes_per_day": recommended_max_adjust_minutes_per_day,
                "warnings": hm_warnings,
            },
            "warnings": warnings,
        }

    # Distribuição exata do TOTAL que precisa ser feito no período restante.
    # Ex.: faltando 16:48 (1008 min) e restam 2 dias => 504 min/dia (08:24).
    total = int(required_total_remaining_minutes)
    if total < 0:
        warnings.append("Você já excedeu o esperado do mês. Sugestão: 00:00 nos dias úteis restantes.")
        total = 0

    base = total // n
    rem = total % n
    
    for idx, di in enumerate(available):
        suggested_total = base + (1 if idx < rem else 0)
        adjust = suggested_total - MINUTES_PER_BUSINESS_DAY

        if abs(adjust) > recommended_max_adjust_minutes_per_day:
            hm_warnings.append(
                f"No dia {di.isoformat()}, ajuste sugerido ({minutes_to_hhmm(adjust)}) acima do recomendado "
                f"({minutes_to_hhmm(recommended_max_adjust_minutes_per_day)})."
            )

        schedule.append(
            {
                "date": di.isoformat(),
                "suggested_adjust_minutes": adjust,
                "suggested_total_for_day_minutes": suggested_total,
                "comment": f"Total sugerido: {minutes_to_hhmm(suggested_total)} (ajuste {minutes_to_hhmm(adjust)}).",
            }
        )

    return {
        "strategy": "UNIFORM",
        "available_business_days": n,
        "target": "ZERO_OUT_MONTH",
        "schedule": schedule,
        "healthy_mode": {
            "enabled": True,
            "recommended_max_adjust_minutes_per_day": recommended_max_adjust_minutes_per_day,
            "warnings": hm_warnings,
        },
        "warnings": warnings,
    }


# ----------------------------
# Schemas (Pydantic)
# ----------------------------
class SchemaResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    kind: str = KIND
    version: str = PONTO_SALDO_VERSION
    title: str = "Saldo de Horas (PDF)"
    description: str = (
        "Lê espelho de ponto em PDF, calcula faltante/sobrante e sugere distribuição "
        "uniforme nos dias úteis restantes do mês (universo = PDF)."
    )
    inputs: List[Dict[str, Any]] = Field(
        default_factory=lambda: [
            {
                "name": "pdf",
                "type": "file",
                "accept": [".pdf"],
                "required": True,
                "label": "Espelho de Ponto (PDF)",
            }
        ]
    )
    outputs: List[str] = Field(default_factory=lambda: ["totals", "days", "planner", "insights"])


class SubmitResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    status: str
    kind: str = KIND


# ----------------------------
# Endpoints
# ----------------------------
@router.get("/schema", response_model=SchemaResponse)
def schema() -> SchemaResponse:
    return SchemaResponse()


@router.get("/ui", response_class=HTMLResponse)
def ui(request: Request):
    return templates.TemplateResponse("ponto_saldo/ui.html", {"request": request})


@router.post("/submit", response_model=SubmitResponse)
async def submit(
    request: Request,
    bg: BackgroundTasks,
    pdf: UploadFile = File(...),
):
    if not pdf.filename or not pdf.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=422, detail="Envie um arquivo .pdf válido.")

    pdf_bytes = await pdf.read()
    if not pdf_bytes or len(pdf_bytes) < 100:
        raise HTTPException(status_code=422, detail="PDF vazio ou inválido.")

    user = (request.session or {}).get("user") or {}
    actor = {
        "cpf": user.get("cpf") or user.get("username") or "",
        "nome": user.get("nome") or user.get("name") or "",
        "email": user.get("email") or "",
        "roles": user.get("roles") or [],
        "is_superuser": bool(user.get("is_superuser") or False),
    }

    sub_id = str(uuid4())
    payload = {
        "filename": pdf.filename,
        "planner_as_of": "TODAY",
        "rules": {
            "daily_expected_minutes": MINUTES_PER_BUSINESS_DAY,
            "pdf_is_source_of_truth": True,
            "ponto_facultativo_is_holiday": True,
            "text_only_leave_counts_as_full_day": True,
            "prefer_result_column": True,
            "on_divergence_trust_result": True,
            "planner": {"strategy": "UNIFORM", "rounding": "NONE"},
        },
    }

    insert_submission(
        {
            "id": sub_id,
            "kind": KIND,
            "version": PONTO_SALDO_VERSION,
            "actor_cpf": actor["cpf"] or None,
            "actor_nome": actor["nome"] or None,
            "actor_email": actor["email"] or None,
            "payload": payload,
            "status": "queued",
            "result": None,
            "error": None,
        }
    )
    add_audit(KIND, "submission.created", actor, {"submission_id": sub_id, "filename": pdf.filename})

    bg.add_task(process_submission, sub_id, pdf_bytes, actor)
    logger.info("ponto_saldo queued id=%s filename=%s", sub_id, pdf.filename)
    return SubmitResponse(id=sub_id, status="queued", kind=KIND)


def process_submission(sub_id: str, pdf_bytes: bytes, actor: Dict[str, Any]) -> None:
    """
    Processa em background:
    - extrai texto do PDF
    - parseia dias
    - aplica regras expected/credited/saldo
    - gera planner uniforme a partir de HOJE (sem arredondar)
    - atualiza submission com result ou error
    """
    try:
        update_submission(sub_id, status="running", error=None)
        add_audit(KIND, "submission.running", actor, {"submission_id": sub_id})

        text = extract_text_from_pdf(pdf_bytes)
        parsed = parse_days(text)
        if not parsed:
            raise ValueError("Não foi possível identificar linhas de dias (dd/mm - ...) no PDF.")

        first_date = date.fromisoformat(parsed[0]["date"])
        period_ym = f"{first_date.year:04d}-{first_date.month:02d}"

        days_out: List[Dict[str, Any]] = []
        expected_total = 0
        credited_total = 0

        for d in parsed:
            markers = d.get("markers_from_pdf") or []
            exp, day_type = compute_expected_minutes(d["date"], markers)
            cred, src, notes = compute_credited_minutes(d["pdf_values"], markers)
            bal = cred - exp

            expected_total += exp
            credited_total += cred

            issues: List[str] = []
            # Se for dia útil esperado e não há total (ex.: hoje só com entrada),
            # marcamos como pendente (não necessariamente erro).
            if exp == MINUTES_PER_BUSINESS_DAY and not (d.get("pdf_values") or {}).get("result_hhmm"):
                # se é hoje ou futuro, não marcar MISSING_DATA
                di = date.fromisoformat(d["date"])
                if di < datetime.now(TZ_SP).date():
                    issues.append("MISSING_DATA")
                else:
                    issues.append("PENDING")

            days_out.append(
                {
                    "date": d["date"],
                    "weekday": d["weekday"],
                    "day_type": day_type,
                    "markers_from_pdf": markers,
                    "expected_minutes": exp,
                    "credited_minutes": cred,
                    "balance_minutes": bal,
                    "pdf_values": d["pdf_values"],
                    "time_tokens_count": d.get("time_tokens_count", 0),
                    "calculation_trace": {"credited_source": src, "notes": notes},
                    "issues": issues,
                }
            )

        # =========
        # Totais para "fechar o mês" + saldo acumulado (sobrando) até agora
        # =========
        today = datetime.now(TZ_SP).date()
        expected_month = expected_total

        expected_so_far = 0
        credited_so_far = 0
        for d in days_out:
            di = date.fromisoformat(d["date"])
            issues = d.get("issues") or []
            include = (di < today) or (di == today and ("PENDING" not in issues))
            if include:
                expected_so_far += int(d.get("expected_minutes") or 0)
                credited_so_far += int(d.get("credited_minutes") or 0)

        so_far_net = credited_so_far - expected_so_far
        so_far_missing = max(-so_far_net, 0)
        so_far_extra = max(so_far_net, 0)

        required_remaining = expected_month - credited_so_far  # pode ser negativo
        missing_total = max(required_remaining, 0)
        extra_total = max(-required_remaining, 0)
        net_balance = credited_so_far - expected_month

        planner = build_planner(
            days=days_out,
            required_total_remaining_minutes=required_remaining,
            recommended_max_adjust_minutes_per_day=120,
        )
        
        result: Dict[str, Any] = {
            "meta": {
                "period": period_ym,
                "generated_at": datetime.now(TZ_SP).isoformat(),
                "planner_as_of": datetime.now(TZ_SP).date().isoformat(),
                "workload": {"expected_minutes_per_business_day": MINUTES_PER_BUSINESS_DAY},
                "source": {"type": "pdf"},
            },
            "totals": {
                "expected_minutes": expected_total,
                # "Creditado" aqui é até agora (base para fechar o mês)
                "credited_minutes": credited_so_far,
                "net_balance_minutes": net_balance,
                "missing_minutes_total": missing_total,
                "extra_minutes_total": extra_total,
                # saldo acumulado (o "sobrando" que permite trabalhar menos no fim do mês)
                "so_far_expected_minutes": expected_so_far,
                "so_far_credited_minutes": credited_so_far,
                "so_far_net_balance_minutes": so_far_net,
                "so_far_missing_minutes_total": so_far_missing,
                "so_far_extra_minutes_total": so_far_extra,
                "pdf_footer_totals": {"worked_minutes": None, "justified_minutes": None, "result_minutes": None},
                "consistency": {"footer_totals_found": False, "matches_footer_totals": None, "notes": []},
            },
            "days": days_out,
            "planner": planner,
            "insights": [],
        }

        missing_days = [x["date"] for x in days_out if "MISSING_DATA" in (x.get("issues") or [])]
        if missing_days:
            result["insights"].append(
                {
                    "severity": "WARN",
                    "kind": "PARSING",
                    "message": f"{len(missing_days)} dia(s) útil(eis) sem horas numéricas (Resultado/Trabalhadas/Justificadas).",
                    "details": {"dates": missing_days[:50]},
                }
            )
        if "[PAGE_" in text:
            result["insights"].append(
                {
                    "severity": "WARN",
                    "kind": "PARSING",
                    "message": "O PDF parece conter páginas sem texto extraível; se for escaneado, pode precisar de OCR.",
                }
            )

        update_submission(sub_id, status="done", result=result, error=None)
        add_audit(KIND, "submission.done", actor, {"submission_id": sub_id})
        logger.info("ponto_saldo done id=%s", sub_id)

    except Exception as e:
        logger.exception("ponto_saldo error id=%s", sub_id)
        update_submission(sub_id, status="error", error=str(e), result=None)
        add_audit(KIND, "submission.error", actor, {"submission_id": sub_id, "error": str(e)})


def _owns_submission(row: Dict[str, Any], user: Dict[str, Any]) -> bool:
    cpf = (user.get("cpf") or "").strip()
    email = (user.get("email") or "").strip().lower()
    if cpf and row.get("actor_cpf") and str(row.get("actor_cpf")).strip() == cpf:
        return True
    if (not cpf) and email and row.get("actor_email"):
        return str(row.get("actor_email")).strip().lower() == email
    # fallback: se tiver cpf mas a submissão só gravou e-mail
    if cpf and email and row.get("actor_email"):
        return str(row.get("actor_email")).strip().lower() == email
    return False


@router.get("/submissions")
def submissions(request: Request, limit: int = 50, offset: int = 0):
    """
    Lista submissões do próprio usuário (CPF preferencial; fallback e-mail).
    Evita 500 do db.list_submissions() quando identidade está ausente.
    """
    user = _current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="not authenticated")

    cpf = (user.get("cpf") or "").strip() or None
    email = (user.get("email") or "").strip().lower() or None
    if not cpf and not email:
        raise HTTPException(
            status_code=422,
            detail="Não foi possível identificar o usuário (sem CPF e e-mail). Faça login novamente.",
        )
    try:
        rows = list_submissions(
            kind=KIND,
            actor_cpf=cpf,
            actor_email=None if cpf else email,
            limit=limit,
            offset=offset,
        )
        return {"items": rows, "limit": limit, "offset": offset}
    except Exception as e:
        logger.exception("ponto_saldo list_submissions failed")
        raise HTTPException(status_code=500, detail=f"Falha ao consultar submissões: {e}")


@router.get("/submissions/{submission_id}")
def submission_get(submission_id: str, request: Request):
    row = get_submission(submission_id)
    if not row or row.get("kind") != KIND:
        raise HTTPException(status_code=404, detail="submission not found")
    user = _current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="not authenticated")
    is_superuser = bool(user.get("is_superuser") or False)
    if (not is_superuser) and (not _owns_submission(row, user)):
        raise HTTPException(status_code=403, detail="forbidden")
    return row


@router.post("/submissions/{submission_id}/download")
def submission_download(submission_id: str, format: Literal["json", "xlsx"] = "json"):
    row = get_submission(submission_id)
    if not row or row.get("kind") != KIND:
        raise HTTPException(status_code=404, detail="submission not found")

    if row.get("status") != "done" or not row.get("result"):
        raise HTTPException(status_code=409, detail="submission not ready")

    result = row["result"]


    if format == "json":
        buf = BytesIO(json.dumps(result, ensure_ascii=False, indent=2).encode("utf-8"))
        filename = f"ponto_saldo_{submission_id}.json"
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        return StreamingResponse(buf, media_type="application/json; charset=utf-8", headers=headers)

    # xlsx
    xlsx_bytes = _xlsx_from_result(result)
    filename = f"ponto_saldo_{submission_id}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
